#!/usr/bin/env python3
"""
Import plots that exist in Tonttilistaus .xlsx (Aihiot sheet, NEW layout:
header row 6, data from row 7) but are missing from the Supabase plots table.

New-layout columns:
    A=Project  B=Priority  C=City  D=Wood/Concrete  E=Address  F=Owner
    G=cityplan  H=in development/valid  I=bu-right sqm  J=comments
    K=Vastuu hlö (skipped)  L=Tarjottu  M..=extra notes

Matching against existing plots is by name, casefolded and whitespace-
collapsed (the DB names came from an identical cleaning of an older copy
of this same sheet).

The `material` column (Puu/Betoni) is filled only if it exists in the DB.

Usage (from Desktop/Antigravity):
    python3 import_new_plots.py            # writes to Supabase
    python3 import_new_plots.py --dry-run  # show what would be inserted
"""

import os
import re
import sys
import json
import uuid
import time
import argparse
import datetime
import requests
import openpyxl
from pathlib import Path

DEFAULT_XLSX = Path.home() / "Downloads" / "Tonttilistaus .xlsx"
ENV_PATH = Path(__file__).parent / "tonttihaku" / ".env.local"
SHEET_NAME = "Aihiot"
DATA_START = 7
TODAY = datetime.date.today().isoformat()
NOW = datetime.datetime.now().isoformat()

# plots."Wood" boolean: true = Puu, false = Betoni, null = unknown
WOOD_MAP = {"wood": True, "puu": True, "betoni": False, "concrete": False}
LABELS = {True: "Puu", False: "Betoni", None: "-"}


def load_env() -> dict:
    env = {}
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    env.update(os.environ)
    return env


def clean_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = s.replace("\n", " ").replace("\t", " ").replace("\xa0", " ")
    s = re.sub(r" {2,}", " ", s)
    if s in ("#VALUE!", "#REF!", "#NAME?", "#N/A"):
        return ""
    return s


def norm_name(name: str) -> str:
    return re.sub(r"\s+", " ", clean_str(name)).casefold().strip()


def parse_kem(value):
    """Returns (buildingRight, is_compound). Same rules as import_plots.py."""
    if value is None:
        return 0, False
    s = clean_str(value).replace(" ", "").replace("?", "")
    if not s:
        return 0, False
    if "+" in s:
        parts = re.findall(r"\d+", s)
        return sum(int(p) for p in parts), True
    if "-" in s:
        parts = re.findall(r"\d+", s)
        if parts:
            return int(parts[-1]), False
    try:
        return int(float(s)), False
    except ValueError:
        nums = re.findall(r"\d+", s)
        return (int(nums[0]) if nums else 0), False


def slug_id(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[^a-zäöåa-z0-9]+", "-", s)
    s = s.strip("-")[:60]
    return s or str(uuid.uuid4())[:8]


def geocode(query: str):
    """Returns (lat, lng, resolved_city) or (0, 0, '')."""
    if not query.strip():
        return 0.0, 0.0, ""
    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": query, "format": "json", "limit": 1,
                    "countrycodes": "fi", "addressdetails": 1},
            headers={"User-Agent": "nordevo-tonttihaku-import/1.0"},
            timeout=10,
        )
        results = resp.json()
        time.sleep(1.1)  # Nominatim rate limit
        if results:
            r = results[0]
            addr = r.get("address", {})
            city = (addr.get("city") or addr.get("town")
                    or addr.get("municipality") or addr.get("county") or "")
            return float(r["lat"]), float(r["lon"]), city
    except Exception as e:
        print(f"  [geocode] WARNING for '{query}': {e}")
    return 0.0, 0.0, ""


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    env = load_env()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL", "")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        sys.exit("ERROR: Supabase env vars not found.")
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    # Does the "Wood" column exist?
    has_wood = requests.get(
        f'{url}/rest/v1/plots?select=Wood&limit=1', headers=headers, timeout=20
    ).status_code == 200
    print(f'"Wood" column in DB: {"yes" if has_wood else "NO — imported without material"}')

    resp = requests.get(f"{url}/rest/v1/plots?select=id,name", headers=headers, timeout=20)
    if resp.status_code != 200:
        sys.exit(f"ERROR fetching plots: {resp.status_code} {resp.text}")
    plots = resp.json()
    existing_names = {norm_name(p["name"]) for p in plots}
    existing_ids = {p["id"] for p in plots}
    print(f"Supabase: {len(plots)} existing plots")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[SHEET_NAME]

    imported, skipped, errors = [], 0, []
    for row_idx in range(DATA_START, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 17)]
        name = clean_str(row[0])
        if not name:
            continue
        if norm_name(name) in existing_names:
            skipped += 1
            continue

        try:
            priority = int(row[1]) if row[1] is not None else 0
        except (ValueError, TypeError):
            priority = 0
        city = clean_str(row[2]).title()
        wood = WOOD_MAP.get(clean_str(row[3]).casefold())
        address = clean_str(row[4])
        seller = clean_str(row[5])
        building_right, is_compound = parse_kem(row[8])
        zoning = clean_str(row[6]) or "AK"
        if is_compound and "KL" not in zoning.upper():
            zoning = "AK+KL"
        kaava_state = clean_str(row[7])   # 'valid' / 'development'
        desc = clean_str(row[9])          # comments column

        # L (idx 11) – Tarjottu: numeric -> offered
        status, offer_price = "Vapaa", 0
        tarjottu = clean_str(row[11])
        if tarjottu:
            try:
                offer_price = int(float(tarjottu))
                status = "Tarjottu"
            except ValueError:
                pass

        # Notes: kaava state + any extra text in columns M..P
        note_texts = []
        if kaava_state:
            note_texts.append(f"Kaava: {kaava_state}")
        for col_i in range(12, 16):
            val = clean_str(row[col_i])
            if val:
                note_texts.append(val)
        notes = [
            {"id": str(uuid.uuid4()), "text": t, "author": "Excel-tuonti",
             "date": TODAY, "timestamp": NOW}
            for t in note_texts
        ]

        # Geocode: address first, then name-based fallbacks
        print(f"  Geocoding [{row_idx}] '{name}'...")
        lat, lng, resolved_city = 0.0, 0.0, ""
        candidates = []
        if address:
            candidates.append(f"{address}, {city}, Finland")
        base = re.sub(r"\b(citycon|tonttikisa|kaupungin)\b", "", name, flags=re.I).strip()
        candidates.append(f"{base}, {city}, Finland" if city else f"{base}, Finland")
        for q in candidates:
            if args.dry_run:
                break  # skip slow network calls on dry runs
            lat, lng, resolved_city = geocode(q)
            if lat:
                break

        plot_id = slug_id(name)
        if plot_id in existing_ids:
            plot_id = f"{plot_id}-2"

        plot = {
            "id": plot_id,
            "name": name,
            "zonings": json.dumps([{"type": zoning, "buildingRight": building_right}], ensure_ascii=False),
            "buildingRight": building_right,
            "area": 0,
            "priceEst": 0,
            "offerPrice": offer_price,
            "desc": desc,
            "seller": seller,
            "status": status,
            "address": address,
            "kunta": resolved_city or city,
            "lat": lat,
            "lng": lng,
            "priority": priority,
            "notes": json.dumps(notes, ensure_ascii=False),
            "createdAt": TODAY,
            "createdBy": "Excel-tuonti",
            "updatedAt": "",
            "updatedBy": "",
            "deadline": "",
            "kiinteistotunnus": "",
            "buyer": "",
            "finalPrice": 0,
            "soldDate": "",
            "offerDate": "",
            "offerDesc": "",
            "contactPerson": "",
            "contactPhone": "",
            "contactEmail": "",
            "contacts": "[]",
            "contactPersons": "[]",
        }
        if has_wood:
            plot["Wood"] = wood

        if args.dry_run:
            imported.append(f"[dry-run] '{name}' (prio={priority}, kunta={plot['kunta']}, "
                            f"kem={building_right}, material={LABELS[wood]})")
            continue

        r = requests.post(f"{url}/rest/v1/plots", headers=headers, json=plot, timeout=20)
        if r.status_code in (200, 201):
            geo = "OK" if lat else "NO GEOCODE — place manually on the map"
            imported.append(f"'{name}' (kunta={plot['kunta']}, kem={building_right}, "
                            f"material={LABELS[wood]}, geo={geo})")
            existing_names.add(norm_name(name))
            existing_ids.add(plot_id)
        else:
            errors.append(f"'{name}': {r.status_code} {r.text[:200]}")

    print(f"\nInserted: {len(imported)}  (skipped {skipped} already in DB)")
    for line in imported:
        print(f"  {line}")
    if errors:
        print(f"Errors: {len(errors)}")
        for e in errors:
            print(f"  {e}")


if __name__ == "__main__":
    main()
