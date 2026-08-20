#!/usr/bin/env python3
"""
Migration script: Tonttilistaus.xlsx (Aihiot sheet) → Supabase plots table
Run from the repo root:  python3 import_plots.py

Requirements:
    pip install openpyxl requests python-dotenv
"""

import os
import re
import sys
import json
import uuid
import time
import datetime
import requests
import openpyxl
from pathlib import Path

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
XLSX_PATH = Path(__file__).parent / "Tonttilistaus.xlsx"
ENV_PATH  = Path(__file__).parent / "tonttihaku" / ".env.local"
SHEET_NAME = "Aihiot"
HEADER_ROW = 5   # 1-indexed
DATA_START  = 6  # 1-indexed
TODAY = datetime.date.today().isoformat()

# ---------------------------------------------------------------------------
# Load environment
# ---------------------------------------------------------------------------
def load_env() -> dict:
    env = {}
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    env.update(os.environ)  # real env always wins
    return env

ENV = load_env()
SUPABASE_URL = ENV.get("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY = ENV.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    sys.exit("ERROR: NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY not found.")

SUPABASE_HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

# ---------------------------------------------------------------------------
# Geocoding (Nominatim — no API key needed, 1 req/s limit)
# ---------------------------------------------------------------------------
def geocode(address: str, city: str) -> tuple[float, float, str]:
    """Returns (lat, lng, resolved_city). Falls back to (0, 0, city)."""
    query = f"{address}, {city}, Finland" if address and city else (address or city or "")
    if not query.strip():
        return 0.0, 0.0, city

    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": query, "format": "json", "limit": 1, "countrycodes": "fi", "addressdetails": 1},
            headers={"User-Agent": "nordevo-tonttihaku-import/1.0"},
            timeout=10,
        )
        results = resp.json()
        time.sleep(1.1)  # respect 1 req/s limit
        if results:
            r = results[0]
            addr = r.get("address", {})
            # Extract municipality from structured fields — never from raw display_name
            resolved_city = (
                addr.get("city") or
                addr.get("town") or
                addr.get("municipality") or
                addr.get("county") or
                city or ""
            )
            # Clear bad pass-through values
            if resolved_city in ("#VALUE!", "#REF!"):
                resolved_city = ""
            return float(r["lat"]), float(r["lon"]), resolved_city
    except Exception as e:
        print(f"  [geocode] WARNING: {e}")

    return 0.0, 0.0, city if city not in ("", "#VALUE!") else ""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def clean_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = s.replace("\n", " ").replace("\t", " ")
    s = re.sub(r" {2,}", " ", s)
    if s in ("#VALUE!", "#REF!", "#NAME?", "#N/A"):
        return ""
    return s


def parse_kem(value) -> tuple[int, bool]:
    """
    Parse k-m2 value. Returns (buildingRight, is_compound_ak_kl).
    - Numbers: return as int
    - "a+b": sum, mark as ak+kl
    - "a-b": take upper value
    - "a?": strip ? and parse
    """
    if value is None:
        return 0, False
    s = clean_str(value).replace(" ", "").replace("?", "")
    if not s:
        return 0, False

    if "+" in s:
        parts = re.findall(r"\d+", s)
        total = sum(int(p) for p in parts)
        return total, True

    if "-" in s:
        parts = re.findall(r"\d+", s)
        if parts:
            return int(parts[-1]), False  # take upper bound

    try:
        return int(float(s)), False
    except ValueError:
        nums = re.findall(r"\d+", s)
        return int(nums[0]) if nums else 0, False


def parse_zoning(value, is_compound: bool) -> str:
    z = clean_str(value)
    if not z:
        z = "AK"
    if is_compound and "KL" not in z.upper():
        z = "AK+KL"
    return z


def build_notes(note_texts: list[str]) -> str:
    notes = []
    for text in note_texts:
        text = clean_str(text)
        if text:
            notes.append({
                "id": str(uuid.uuid4()),
                "text": text,
                "author": "Excel-tuonti",
                "date": TODAY,
            })
    return json.dumps(notes, ensure_ascii=False)


def slug_id(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[^a-zäöåa-z0-9]+", "-", s)
    s = s.strip("-")[:60]
    return s or str(uuid.uuid4())[:8]


# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------
def get_existing_names() -> set:
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/plots?select=name",
        headers=SUPABASE_HEADERS,
        timeout=15,
    )
    if resp.status_code == 200:
        return {row["name"].strip() for row in resp.json()}
    print(f"WARNING: could not fetch existing plots ({resp.status_code}): {resp.text}")
    return set()


def insert_plot(plot: dict) -> tuple[bool, str]:
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/plots",
        headers=SUPABASE_HEADERS,
        json=plot,
        timeout=15,
    )
    if resp.status_code in (200, 201):
        return True, ""
    return False, resp.text


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print(f"Opening {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    print(f"Fetching existing plot names from Supabase...")
    existing_names = get_existing_names()
    print(f"  Found {len(existing_names)} existing plots.\n")

    imported  = []
    skipped   = []
    errors    = []
    flagged   = []  # rows needing manual review

    max_row = ws.max_row
    for row_idx in range(DATA_START, max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 17)]

        # A – Hanke (name)
        name = clean_str(row[0])
        if not name:
            continue  # blank row

        # B – Prioriteetti
        try:
            priority = int(row[1]) if row[1] is not None else 3
        except (ValueError, TypeError):
            priority = 3

        # C – Kaupunki
        raw_city = clean_str(row[2])  # may be "" after cleaning #VALUE!

        # D – Osoite
        address = clean_str(row[3])

        # E – Omistaja (seller)
        seller = clean_str(row[4])

        # F – kaava
        # G – k-m2
        kem_raw = row[6]
        building_right, is_compound = parse_kem(kem_raw)
        if isinstance(kem_raw, str) and "-" in kem_raw:
            flagged.append(f"Row {row_idx} '{name}': k-m2 range '{kem_raw}' → took upper value {building_right}")

        zoning = parse_zoning(row[5], is_compound)

        # H – Hankkeen tilanne → desc
        desc = clean_str(row[7])

        # I – Vastuu hlö → SKIPPED (internal only)

        # J – Tarjottu → status + offerPrice
        tarjottu = row[9]
        status = "Vapaa"
        offer_price = 0
        if tarjottu is not None:
            try:
                offer_price = int(float(str(tarjottu)))
                status = "Tarjottu"
            except (ValueError, TypeError):
                pass

        # K (col 10) + extra cols L–P (11–15) → notes
        note_texts = []
        for col_i in range(10, 16):
            val = clean_str(row[col_i]) if col_i < len(row) else ""
            if val:
                note_texts.append(val)

        # Duplicate check
        if name in existing_names:
            skipped.append(f"Row {row_idx}: '{name}' already exists — skipped")
            continue

        # Geocode
        print(f"  Geocoding [{row_idx}] '{name}'...")
        lat, lng, resolved_city = geocode(address, raw_city)
        if not resolved_city and not lat:
            # Try geocoding just the name as a fallback
            lat2, lng2, resolved_city2 = geocode(name, "Finland")
            if lat2:
                lat, lng, resolved_city = lat2, lng2, resolved_city2

        plot_id = slug_id(name)

        plot = {
            "id":            plot_id,
            "name":          name,
            "zonings":       zoning,
            "buildingRight": building_right,
            "area":          0,
            "priceEst":      0,
            "offerPrice":    offer_price,
            "desc":          desc,
            "seller":        seller,
            "status":        status,
            "address":       address,
            "kunta":         resolved_city,
            "lat":           lat,
            "lng":           lng,
            "priority":      priority,
            "notes":         build_notes(note_texts),
            "createdAt":     TODAY,
            "createdBy":     "Excel-tuonti",
            "updatedAt":     TODAY,
            "updatedBy":     "Excel-tuonti",
            "deadline":      "",
            "kiinteistotunnus": "",
            "buyer":         "",
            "finalPrice":    0,
            "soldDate":      "",
            "offerDate":     "",
            "offerDesc":     "",
            "contactPerson": "",
            "contactPhone":  "",
            "contactEmail":  "",
            "contacts":      "[]",
            "contactPersons": "[]",
        }

        ok, err = insert_plot(plot)
        if ok:
            imported.append(f"Row {row_idx}: '{name}' (status={status}, kem={building_right}, city={resolved_city})")
            existing_names.add(name)
        else:
            errors.append(f"Row {row_idx}: '{name}' INSERT ERROR: {err}")

    # ---------------------------------------------------------------------------
    # Report
    # ---------------------------------------------------------------------------
    report_lines = [
        "=" * 70,
        f"MIGRATION REPORT — {TODAY}",
        "=" * 70,
        f"Imported:  {len(imported)}",
        f"Skipped (duplicates):  {len(skipped)}",
        f"Errors:    {len(errors)}",
        f"Flagged for review:  {len(flagged)}",
        "",
    ]

    if imported:
        report_lines += ["--- IMPORTED ---"]
        report_lines += imported
        report_lines += [""]

    if skipped:
        report_lines += ["--- SKIPPED (DUPLICATES) ---"]
        report_lines += skipped
        report_lines += [""]

    if flagged:
        report_lines += ["--- FLAGGED FOR MANUAL REVIEW ---"]
        report_lines += flagged
        report_lines += [""]

    if errors:
        report_lines += ["--- ERRORS ---"]
        report_lines += errors
        report_lines += [""]

    report_text = "\n".join(report_lines)
    report_path = Path(__file__).parent / "migration_report.txt"
    report_path.write_text(report_text, encoding="utf-8")

    print("\n" + report_text)
    print(f"\nReport saved to: {report_path}")


if __name__ == "__main__":
    main()
