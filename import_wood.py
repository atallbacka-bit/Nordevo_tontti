#!/usr/bin/env python3
"""
Import wood/concrete potential from Tonttilistaus .xlsx (Aihiot sheet,
"Wood /Concrete" column D) into the Supabase plots."Wood" boolean column:
    true = Wood (Puu), false = Betoni, null = unknown (blank cell).

Matches Excel rows to existing plots by name (same cleaning rules as the
original import_plots.py migration).

Usage (from Desktop/Antigravity):
    python3 import_wood.py            # writes to Supabase
    python3 import_wood.py --dry-run  # only shows what would be updated
    python3 import_wood.py --xlsx "/path/to/Tonttilistaus .xlsx"
"""

import os
import re
import sys
import argparse
import requests
import openpyxl
from pathlib import Path

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DEFAULT_XLSX = Path.home() / "Downloads" / "Tonttilistaus .xlsx"
ENV_PATH = Path(__file__).parent / "tonttihaku" / ".env.local"
SHEET_NAME = "Aihiot"
DATA_START = 7  # 1-indexed; header is on row 6 in the new layout
NAME_COL = 1    # A - Project
MATERIAL_COL = 4  # D - "Wood /Concrete"

WOOD_MAP = {
    "wood": True,
    "puu": True,
    "betoni": False,
    "concrete": False,
}

LABELS = {True: "Puu", False: "Betoni"}


def load_env() -> dict:
    env = {}
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    env.update(os.environ)
    return env


def clean_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = s.replace("\n", " ").replace("\t", " ").replace("\xa0", " ")
    s = re.sub(r" {2,}", " ", s)
    if s in ("#VALUE!", "#REF!", "#NAME?", "#N/A"):
        return ""
    return s


def norm_name(name: str) -> str:
    """Casefolded, whitespace-collapsed key for name matching."""
    return re.sub(r"\s+", " ", clean_str(name)).casefold().strip()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Report matches without writing")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Path to Tonttilistaus xlsx")
    args = parser.parse_args()

    env = load_env()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL", "")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        sys.exit("ERROR: NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY not found.")
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }

    if not args.xlsx.exists():
        sys.exit(f"ERROR: Excel file not found: {args.xlsx}")

    # -----------------------------------------------------------------------
    # 1. Read wood/concrete data from Excel
    # -----------------------------------------------------------------------
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[SHEET_NAME]

    excel_wood: dict[str, tuple[str, bool]] = {}  # norm_name -> (display_name, wood)
    conflicts = []
    for row_idx in range(DATA_START, ws.max_row + 1):
        name = clean_str(ws.cell(row=row_idx, column=NAME_COL).value)
        if not name:
            continue
        raw = clean_str(ws.cell(row=row_idx, column=MATERIAL_COL).value).casefold()
        if not raw:
            continue  # blank cell: leave existing value untouched
        if raw not in WOOD_MAP:
            print(f"  WARNING row {row_idx} '{name}': unrecognized material '{raw}' — skipped")
            continue
        wood = WOOD_MAP[raw]
        k = norm_name(name)
        if k in excel_wood and excel_wood[k][1] != wood:
            conflicts.append(f"'{name}': {LABELS[excel_wood[k][1]]} vs {LABELS[wood]} — using {LABELS[wood]}")
        excel_wood[k] = (name, wood)

    n_puu = sum(1 for _, w in excel_wood.values() if w is True)
    n_betoni = sum(1 for _, w in excel_wood.values() if w is False)
    print(f"Excel: {len(excel_wood)} projects with material ({n_puu} Puu, {n_betoni} Betoni)")
    for c in conflicts:
        print(f"  CONFLICT: {c}")

    # -----------------------------------------------------------------------
    # 2. Fetch existing plots (and verify the "Wood" column exists)
    # -----------------------------------------------------------------------
    resp = requests.get(f'{url}/rest/v1/plots?select=id,name,Wood', headers=headers, timeout=20)
    if resp.status_code != 200 and "Wood" in resp.text:
        msg = (
            'The plots."Wood" column does not exist yet.\n'
            "Run this in the Supabase SQL Editor, then re-run this script:\n\n"
            '    ALTER TABLE plots ADD COLUMN IF NOT EXISTS "Wood" boolean;\n'
        )
        if not args.dry_run:
            sys.exit("ERROR: " + msg)
        print("NOTE: " + msg)
        resp = requests.get(f"{url}/rest/v1/plots?select=id,name", headers=headers, timeout=20)
    if resp.status_code != 200:
        sys.exit(f"ERROR fetching plots ({resp.status_code}): {resp.text}")
    plots = resp.json()
    print(f"Supabase: {len(plots)} plots")

    by_name: dict[str, list[dict]] = {}
    for p in plots:
        by_name.setdefault(norm_name(p.get("name", "")), []).append(p)

    # -----------------------------------------------------------------------
    # 3. Match and update
    # -----------------------------------------------------------------------
    updated, unchanged, unmatched = [], [], []
    for k, (display, wood) in sorted(excel_wood.items()):
        targets = by_name.get(k)
        if not targets:
            unmatched.append(f"'{display}' ({LABELS[wood]})")
            continue
        for plot in targets:
            if plot.get("Wood") is wood:
                unchanged.append(f"'{plot['name']}' already {LABELS[wood]}")
                continue
            if args.dry_run:
                updated.append(f"'{plot['name']}' -> {LABELS[wood]}  [dry-run]")
                continue
            r = requests.patch(
                f"{url}/rest/v1/plots",
                params={"id": f"eq.{plot['id']}"},
                headers=headers,
                json={"Wood": wood},
                timeout=20,
            )
            if r.status_code in (200, 204):
                updated.append(f"'{plot['name']}' -> {LABELS[wood]}")
            else:
                print(f"  ERROR updating '{plot['name']}' ({r.status_code}): {r.text}")

    # -----------------------------------------------------------------------
    # 4. Report
    # -----------------------------------------------------------------------
    print()
    print(f"{'Would update' if args.dry_run else 'Updated'}: {len(updated)}")
    for u in updated:
        print(f"  {u}")
    if unchanged:
        print(f"Already correct: {len(unchanged)}")
    if unmatched:
        print(f"No matching plot in Supabase: {len(unmatched)}")
        for u in unmatched:
            print(f"  {u}")


if __name__ == "__main__":
    main()
